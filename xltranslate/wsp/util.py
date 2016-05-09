#

import logging

import openpyxl

log = logging.getLogger(__name__)


def sanitise_string(text):
    text = u'%s' % (text, )
    text = text.strip()
    text = text.replace('\n', ' ')
    text = text.encode('ascii', 'ignore')
    return text


def get_tables(sheet, table_meta_list):
    table_meta_dict = _to_table_meta_dict(table_meta_list)
    row_gen = sheet.get_squared_range(1, 1, sheet.max_column, sheet.max_row)
    cells = [row for row in row_gen]
    table_start_rows = _get_table_start_rows(cells, table_meta_list)
    tables = {}
    for index in range(0, len(table_start_rows)):
        tname, start_row = table_start_rows[index]
        first_data_row_number = table_meta_dict[tname]["first-data-row"]
        end_row = sheet.max_row
        next_index = index + 1
        if next_index < len(table_start_rows):
            end_row = table_start_rows[next_index][1]
        tables[tname] = _get_table_data_cells(cells, start_row, end_row,
                                              first_data_row_number)
    return tables


def _to_table_meta_dict(table_meta_list):
    d = {}
    for tmeta in table_meta_list:
        d[tmeta["name"]] = tmeta
    return d


def _get_table_start_rows(cells, table_meta_list):
    start_rows = []
    for table_meta in table_meta_list:
        tname = table_meta["name"]
        start_row = _get_starting_row(cells, tname)
        if start_row is None:
            log.info("Could not locate table '%s'. Ignoring.", tname)
            continue
        start_rows.append((tname, start_row))
    sorted_start_rows = sorted(start_rows, key=lambda x: x[1])
    return sorted_start_rows


def _get_starting_row(cells, tname):
    for row, row_number in zip(cells, range(0, len(cells))):
        c1_value = sanitise_string(row[0].value)
        if c1_value.startswith(tname):
            return row_number
    return None


def _get_table_data_cells(cells, start_row, end_row, first_data_row_number):
    table = cells[start_row:end_row]
    empty = 0
    first_data_row = table[first_data_row_number]
    for cell in reversed(first_data_row):
        if cell.value is None:
            empty += 1
        else:
            break
    last_index = len(first_data_row) - empty
    ntable = []
    log.debug("Trimming extracted table of empty values. Row-len=%d, empty=%d",
              len(first_data_row), empty)
    for row in table:
        ntable.append(row[0:last_index])
    return ntable


class TypeATable(object):
    def __init__(self, table_data):
        self._table_data = table_data
        self._row_len = len(table_data)
        self._col_len = len(table_data[0])
        self._variable_names = []
        self._variable_rows = []
        self._data_set = []
        self._find_variable_info()
        self._extract_data()

    @property
    def variables(self):
        return self._variable_names

    @property
    def data_set(self):
        return self._data_set

    def _find_variable_info(self):
        for row, index in zip(self._table_data, range(0, self._row_len)):
            col1_data = sanitise_string(row[0].value)
            col2_data = row[1].value
            if col1_data and col2_data:
                self._variable_names.append(col1_data)
                self._variable_rows.append(index)

    def _extract_data(self):
        self._data_set = []
        for col in range(1, self._col_len):
            self._data_set.append(self._extract_data_for_column(col))

    def _extract_data_for_column(self, col):
        data_set = []
        for row_no in self._variable_rows:
            cell = self._table_data[row_no][col]
            data = cell.value
            if cell.data_type == openpyxl.cell.Cell.TYPE_STRING:
                data = sanitise_string(data)
            if cell.is_date:
                data = data.strftime("%b-%d-%Y")
            if data in ('-', ):
                data = 0
            data = "%s" % (data, )
            data_set.append(data)
        return data_set

    def dump_to_screen(self):
        # construct a format-line for pretty printing
        fmt_list = []
        for col in range(0, len(self._variable_names)):
            col_size = len(self._variable_names[col])
            for row in range(0, len(self._data_set)):
                val = "%s" % (self._data_set[row][col], )
                if val:
                    data_sz = len(val)
                    if data_sz > col_size:
                        col_size = data_sz
            fmt_list.append("{:<%d}" % (col_size, ))
        fmt_line = u' '.join(fmt_list)
        # dump on screen
        print(fmt_line.format(*self._variable_names))
        for ds_line in self._data_set:
            fmtted = fmt_line.format(*ds_line)
            print(fmtted)


def create_type_a_table(raw_table):
    if len(raw_table) < 2:
        return None
    if len(raw_table[0]) < 2:
        return None
    return TypeATable(raw_table)


def _dump_to_hdf5(variables, data_set, h5_group, dataset_name):
    max_string_length = _compute_max_string_length(data_set)
    dtype = "S%d" % (max_string_length + 1, )
    col_size = len(data_set[0])
    row_size = len(data_set)
    dataset_name = dataset_name.replace('/', ' ')
    h5dset = h5_group.create_dataset(dataset_name, (row_size, col_size),
                                     dtype=dtype)
    for row in range(0, row_size):
        h5dset[row] = data_set[row]
    h5dset.attrs["variables"] = variables


def to_hdf5(group, wspobj):
    for k, v in wspobj.metadata.items():
        group.attrs[k] = v
    for name, table in wspobj.tables.items():
        _dump_to_hdf5(table.variables, table.data_set, group, name)


def _compute_max_string_length(data_set):
    col_count = len(data_set[0])
    max_length = 0
    for row in data_set:
        for val in row:
            val_len = len(val)
            if val_len > max_length:
                max_length = val_len
    return max_length


def get_sheet_metadata(sheet):
    row_gen = sheet.get_squared_range(1, 1, sheet.max_column, sheet.max_row)
    cells = [row for row in row_gen]
    set1 = _extract_metadata(cells, 6, 1)
    set2 = _extract_metadata(cells, 6, 4)
    set1.update(set2)
    return set1

def _extract_metadata(cells, start_row, start_col):
    """The start_row and start_col are indexed from 0 in the 2-D array,
    cells

    """

    row_no = start_row
    col_no = start_col
    row_count = len(cells)
    data = {}
    while True:
        k_cell = cells[row_no][col_no]
        if k_cell.value is None:
            return data
        v_cell = cells[row_no][col_no + 1]
        if v_cell.value is None:
            return data
        k = sanitise_string(k_cell.value)
        v = sanitise_string(v_cell.value)
        if k.endswith(":"):
            k = k[:-1]
        data[k] = v
        row_no += 1
        if row_no == row_count:
            log.error("Metadata table cannot go all the way to the end"
                      " of the sheet. Something wrong. Bailing.")
            return {}
