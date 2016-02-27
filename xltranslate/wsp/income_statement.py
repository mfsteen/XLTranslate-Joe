#
import logging

import openpyxl

log = logging.getLogger(__name__)


class IncomeStatement(object):
    def __init__(self, sheet):
        self._sheet = sheet
        self._starting_row = 15
        self._variables = self._extract_col_variables()
        self._last_data_column = self._find_last_data_column()
        self._data_set = self._extract_col_data()

    @property
    def variables(self):
        return self._variables

    @property
    def data_set(self):
        return self._data_set

    def _find_last_data_column(self):
        for col in range(2, (self._sheet.max_column + 1)):
            data = self._sheet.cell(column=col, row=self._starting_row).value
            if data is None:
                return (col - 1)
        raise RuntimeError("No convergence on number-of-data-columns")

    def _extract_col_variables(self):
        variables = []
        row_number = self._starting_row
        empty_lines = 0
        while True:
            var = self._sheet.cell(column=1, row=row_number)
            data = self._sheet.cell(column=2, row=row_number)
            if var.value and data.value:
                variables.append((row_number, var.value))
                empty_lines = 0
            else:
                empty_lines += 1
            if (empty_lines >= 5) or (row_number == self._sheet.max_row):
                break
            row_number += 1
        return variables

    def _extract_col_data(self):
        data_set = []
        for col in range(2, (self._last_data_column + 1)):
            data_set.append(self._extract_col_data_per_column(col))
        return data_set

    def _extract_col_data_per_column(self, column_number):
        data_set = []
        for row, var in self._variables:
            cell = self._sheet.cell(column=column_number, row=row)
            data = cell.value
            if cell.data_type == openpyxl.cell.Cell.TYPE_STRING:
                data = data.strip()
                data = data.replace('\n', ' ')
            if data in ('-', ):
                data = 0
            # log.debug("c=%d, r=%d type=%s data=%s" %
            #           (column_number, row, cell.data_type, data))
            data_set.append(data)
        return data_set

    def dump(self):
        var = [var.strip() for row, var in self._variables]
        data_set = []
        data_set.append(var)
        for ds_line in self.data_set:
            ds_line = [str(d) for d in ds_line]
            data_set.append(ds_line)
        # construct a format-line for pretty printing
        fmt_list = []
        for col in range(0, len(var)):
            col_size = 0
            for row in range(0, len(data_set)):
                data_sz = len(data_set[row][col])
                if  data_sz > col_size:
                    col_size = data_sz
            fmt_list.append("{:<%d}" % (col_size, ))
        fmt_line = ' '.join(fmt_list)
        # dump on screen
        for ds_line in data_set:
            print(fmt_line.format(*ds_line))
